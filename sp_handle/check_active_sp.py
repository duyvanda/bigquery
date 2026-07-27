import os
import glob
import sys
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

SP_CACHE_PATH = r'd:\bigquery\cache\sp_jobs_cache.csv'

CREDENTIALS_PATH = 'D:/bigquery1508.json'
PROJECT_ID = 'spatial-vision-343005'
ACTIVE_JOB_FILE = r'd:\bigquery\sp_handle\call_active_job.md'
STAGING_TEMP_DIR = r'd:\bigquery\staging_temp'
EXCEL_OUTPUT_PATH = r'd:\bigquery\danh_sach_store_het_dung.xlsx'

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = CREDENTIALS_PATH

def check_active_stores():
    """
    Rule Active SP = (Nằm trong call_active_job.md) AND (Có thực thi thực tế JOBS 180d).
    Fix: scan toàn bộ query string để bắt multi-statement CALL scripts.
    """
    client = bigquery.Client(project=PROJECT_ID) if not os.path.exists(SP_CACHE_PATH) else None
    
    print("==========================================")
    print("PHÂN TÍCH STORED PROCEDURES - ĐIỀU KIỆN KÉP (MULTI-STATEMENT SCRIPT SAFE)")
    print("Rule: Active SP = Kê khai trong call_active_job.md AND có CALL thực tế trong 180d")
    print("==========================================\n")

    # 1. Đọc 107 SPs từ call_active_job.md
    with open(ACTIVE_JOB_FILE, 'r', encoding='utf-8') as f:
        text = f.read()

    called_sps = set()
    for line in text.split('\n'):
        for c in line.split(';'):
            if 'CALL' in c.upper():
                parts = c.split('.')
                if len(parts) >= 2:
                    clean_name = re.sub(r'[^a-zA-Z0-9_]', '', parts[-1]).lower()
                    if clean_name:
                        called_sps.add(clean_name)

    print(f"1. Số SPs được kê khai trong call_active_job.md: {len(called_sps)}")

    # 2. Đọc SP usage từ cache local (chạy cache\fetch_jobs_cache.py trước nếu chưa có)
    if os.path.exists(SP_CACHE_PATH):
        print(f"\n2. Đọc SP JOBS cache từ local: {SP_CACHE_PATH}")
        df_cache = pd.read_csv(SP_CACHE_PATH, parse_dates=['last_used_time'])
        # Filter lại 180 ngày tính từ NOW (cache có thể cũ hơn ngày fetch)
        cutoff = pd.Timestamp.now(tz='UTC') - pd.Timedelta(days=180)
        df_cache['last_used_time'] = pd.to_datetime(df_cache['last_used_time'], utc=True)
        df_cache = df_cache[df_cache['last_used_time'] >= cutoff]
        print(f"   -> Cache có {len(df_cache)} SP records trong 180 ngày gần nhất.")
    else:
        print(f"\n2. Cache chưa có, đang query BigQuery (chạy cache/fetch_jobs_cache.py để tạo cache)...")
        sql = """
        SELECT name, MAX(last_used_time) AS last_used_time, STRING_AGG(DISTINCT users) AS users
        FROM (
            SELECT sp_name AS name, MAX(creation_time) AS last_used_time, user_email AS users
            FROM `region-asia-southeast1`.INFORMATION_SCHEMA.JOBS,
            UNNEST(REGEXP_EXTRACT_ALL(LOWER(query), r'staging_temp[.`]+([a-z0-9_]+)')) AS sp_name
            WHERE creation_time >= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 180 DAY)
            GROUP BY name, users
        )
        GROUP BY name
        """
        rows = list(client.query(sql).result())
        df_cache = pd.DataFrame([dict(r) for r in rows])
        df_cache.columns = ['name', 'last_used_time', 'users']
        os.makedirs(os.path.dirname(SP_CACHE_PATH), exist_ok=True)
        df_cache.to_csv(SP_CACHE_PATH, index=False, encoding='utf-8-sig')
        print(f"   -> Đã lưu cache: {SP_CACHE_PATH}")

    # 3. Map cache về executed_sps dict
    executed_sps = {}  # sp_name -> last_executed_time
    for _, row in df_cache.iterrows():
        sp = str(row['name']).strip().lower()
        if sp in called_sps:
            executed_sps[sp] = row['last_used_time']

    print(f"   -> Số SPs trong call_active_job.md CÓ CHẠY THỰC TẾ 180D: {len(executed_sps)} / {len(called_sps)}")

    inactive_in_md = called_sps - set(executed_sps.keys())
    if inactive_in_md:
        print(f"   -> Số SPs trong call_active_job.md KHÔNG CHẠY TRONG 180D: {len(inactive_in_md)}")
        for s in sorted(list(inactive_in_md)):
            print(f"      - {s}")

    # 4. Lấy metadata (Created & Last Altered) từ INFORMATION_SCHEMA.ROUTINES
    print("\n3. Đang lấy Metadata từ INFORMATION_SCHEMA.ROUTINES...")
    datasets = ['staging_temp', 'staging', 'warehouse']
    routine_meta = {}
    for ds in datasets:
        try:
            rows = list(client.query(f"""
                SELECT routine_name, created, last_altered
                FROM `{PROJECT_ID}.{ds}.INFORMATION_SCHEMA.ROUTINES`
            """).result())
            for r in rows:
                routine_meta[r.routine_name.lower()] = {
                    'created': r.created.strftime('%Y-%m-%d %H:%M:%S') if r.created else '',
                    'last_altered': r.last_altered.strftime('%Y-%m-%d %H:%M:%S') if r.last_altered else ''
                }
        except Exception:
            pass

    # 5. Phân loại tất cả file .sql
    active_sql = glob.glob(os.path.join(STAGING_TEMP_DIR, 'active', '*.sql'))
    backup_sql = glob.glob(os.path.join(STAGING_TEMP_DIR, 'backup', '*.sql'))
    all_sql = active_sql + backup_sql

    print(f"\n4. Phân loại {len(all_sql)} file SQL SPs ({len(active_sql)} active, {len(backup_sql)} backup)...")

    results = []
    for sf in sorted(all_sql):
        filename = os.path.basename(sf)
        sp_name = os.path.splitext(filename)[0].lower()
        
        in_md = sp_name in called_sps
        job_time = executed_sps.get(sp_name)
        has_jobs = job_time is not None
        
        # ĐIỀU KIỆN KÉP
        is_active = in_md and has_jobs
        
        # Xác định trạng thái cụ thể
        if is_active:
            status = 'Active (Đang chạy định kỳ)'
        elif in_md and not has_jobs:
            status = 'Hết dùng (Có trong MD nhưng 180d không chạy)'
        else:
            status = 'Hết dùng (Không có trong Active Job)'

        meta = routine_meta.get(sp_name, {'created': 'N/A', 'last_altered': 'N/A'})
        last_used_str = job_time.strftime('%Y-%m-%d %H:%M:%S') if has_jobs else 'Không gọi trong 180d'
        
        results.append({
            'Tên Stored Procedure': sp_name,
            'Tên File SQL': filename,
            'Trong call_active_job.md': 'Có' if in_md else 'Không',
            'Chạy thực tế 180d': 'Có' if has_jobs else 'Không',
            'Lần cuối chạy (JOBS 180d)': last_used_str,
            'Ngày tạo (Created)': meta['created'],
            'Cập nhật lần cuối (Last Altered)': meta['last_altered'],
            'Kích thước File (KB)': round(os.path.getsize(sf) / 1024, 2),
            'Thư mục chứa': 'active/' if 'active' in sf else 'backup/',
            'Trạng thái': status
        })

    df = pd.DataFrame(results)
    df_active = df[df['Trạng thái'].str.startswith('Active')]
    df_unused = df[~df['Trạng thái'].str.startswith('Active')]
    df_unused = df_unused.reset_index(drop=True)
    df_unused.insert(0, 'STT', range(1, len(df_unused) + 1))

    print(f"\n================ TỔNG HỢP STORED PROCEDURES ================")
    print(f"Tổng số Stored Procedures local: {len(df)}")
    print(f"  - Active SPs THỰC SỰ (Điều kiện kép): {len(df_active)}")
    print(f"  - Unused SPs (Hết dùng): {len(df_unused)}")
    print("=============================================================\n")

    # 6. Cập nhật thư mục active/backup nếu phân loại sai
    moved = 0
    st_active_dir = os.path.join(STAGING_TEMP_DIR, 'active')
    st_backup_dir = os.path.join(STAGING_TEMP_DIR, 'backup')
    import shutil
    for sf in all_sql:
        filename = os.path.basename(sf)
        sp_name = os.path.splitext(filename)[0].lower()
        in_md = sp_name in called_sps
        has_jobs = sp_name in executed_sps
        is_active = in_md and has_jobs
        
        if is_active and 'backup' in sf:
            shutil.move(sf, os.path.join(st_active_dir, filename))
            print(f"[MOVE -> active/] {filename}")
            moved += 1
        elif not is_active and 'active' in sf:
            shutil.move(sf, os.path.join(st_backup_dir, filename))
            print(f"[MOVE -> backup/] {filename}")
            moved += 1

    if moved > 0:
        print(f"\n-> Đã điều chỉnh {moved} file vào đúng thư mục.")
    else:
        print("-> Thư mục active/backup đã phân loại chính xác.")

    # 7. Xuất Excel
    out_file = EXCEL_OUTPUT_PATH
    try:
        if os.path.exists(out_file):
            os.remove(out_file)
    except Exception:
        out_file = r'd:\bigquery\danh_sach_store_het_dung_moi.xlsx'

    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        df_unused.to_excel(writer, index=False, sheet_name='Store Hết Dùng')
        workbook = writer.book
        worksheet = writer.sheets['Store Hết Dùng']
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        for col_num in range(1, len(df_unused.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        worksheet.row_dimensions[1].height = 28

        data_font = Font(name='Segoe UI', size=10)
        zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        for row_num in range(2, len(df_unused) + 2):
            worksheet.row_dimensions[row_num].height = 20
            for col_num in range(1, len(df_unused.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if row_num % 2 == 1:
                    cell.fill = zebra_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_num in [2, 7]:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            worksheet.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

    print(f"\n[+] Đã xuất báo cáo Excel Store Hết Dùng tại: {out_file}")

if __name__ == '__main__':
    check_active_stores()
