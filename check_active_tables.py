import os
import glob
import re
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

CREDENTIALS_PATH = 'D:/bigquery1508.json'
PROJECT_ID = 'spatial-vision-343005'
DATASET_ID = 'staging'
EXCEL_REPORT_PATH = r'd:\bigquery\danh_sach_table_staging_usage.xlsx'

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = CREDENTIALS_PATH
client = bigquery.Client(project=PROJECT_ID)

def analyze_staging_tables():
    print("==========================================")
    print("PHÂN TÍCH USAGE TẤT CẢ TABLES TRONG DATASET STAGING")
    print("==========================================\n")

    # 1. Truy vấn Server-side Metadata (Rows, Bytes, Last Modified) từ INFORMATION_SCHEMA.TABLE_STORAGE
    print(f"1. Đang truy vấn Server-side Metadata Tables từ BigQuery Dataset '{DATASET_ID}'...")
    query_storage = f"""
    SELECT 
        table_name,
        total_rows,
        total_logical_bytes,
        storage_last_modified_time
    FROM `region-asia-southeast1`.INFORMATION_SCHEMA.TABLE_STORAGE
    WHERE table_schema = '{DATASET_ID}'
    """
    
    table_meta = {}
    try:
        rows = list(client.query(query_storage).result())
        for r in rows:
            table_meta[r.table_name] = {
                'total_rows': r.total_rows or 0,
                'total_bytes': r.total_logical_bytes or 0,
                'last_modified': r.storage_last_modified_time
            }
        print(f"-> Ghi nhận Metadata cho {len(table_meta)} Tables trong dataset '{DATASET_ID}'.")
    except Exception as e:
        print(f"[!] Lỗi truy vấn TABLE_STORAGE: {e}")

    # Fallback nếu TABLE_STORAGE không có dữ liệu -> list_tables
    if not table_meta:
        tables_list = list(client.list_tables(DATASET_ID))
        for t in tables_list:
            table_meta[t.table_id] = {'total_rows': 0, 'total_bytes': 0, 'last_modified': None}

    table_names = sorted(list(table_meta.keys()))

    # 2. Truy vấn Server-side lịch sử Query (JOBS - 180 ngày) cho Dataset 'staging'
    print("\n2. Đang truy vấn Server-side lịch sử Query (JOBS - 180 ngày) từ BigQuery...")
    query_jobs = f"""
    SELECT 
        REGEXP_EXTRACT(query, r'(?i)(?:{DATASET_ID}|`{DATASET_ID}`)\.(`?[\w]+`?)') AS raw_tbl,
        MAX(creation_time) AS last_queried_time,
        COUNT(1) AS total_queries,
        STRING_AGG(DISTINCT user_email, ', ') AS user_emails
    FROM `region-asia-southeast1`.INFORMATION_SCHEMA.JOBS
    WHERE creation_time >= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 180 DAY)
      AND (LOWER(query) LIKE '%{DATASET_ID}%' OR LOWER(query) LIKE '%table%')
    GROUP BY raw_tbl
    HAVING raw_tbl IS NOT NULL
    """
    
    table_job_usage = {}
    try:
        for row in client.query(query_jobs).result():
            clean_tbl = str(row.raw_tbl).replace('`', '').strip()
            table_job_usage[clean_tbl] = {
                'last_queried': row.last_queried_time,
                'total_queries': row.total_queries,
                'user_emails': row.user_emails or ''
            }
        print(f"-> Ghi nhận {len(table_job_usage)} Tables có lịch sử truy vấn trong 180 ngày qua.")
    except Exception as e:
        print(f"[!] Lỗi truy vấn JOBS: {e}")

    # 3. Quét tham chiếu trong Code Stored Procedures (.sql) & Views (.sql)
    print("\n3. Đang quét tham chiếu Table trong Code SPs (staging_temp) & Views (warehouse)...")
    sp_files = glob.glob(r'd:\bigquery\staging_temp\*.sql')
    view_files = glob.glob(r'd:\bigquery\warehouse\*.sql')
    
    referenced_in_code = {}

    for sf in sp_files:
        sp_name = os.path.basename(sf)
        with open(sf, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
        for tn in table_names:
            tn_lower = tn.lower()
            if f"staging.{tn_lower}" in content or f"`staging`.`{tn_lower}`" in content or f"`{tn_lower}`" in content:
                referenced_in_code.setdefault(tn, set()).add(f"SP: {sp_name}")

    for vf in view_files:
        v_name = os.path.basename(vf)
        with open(vf, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
        for tn in table_names:
            tn_lower = tn.lower()
            if f"staging.{tn_lower}" in content or f"`staging`.`{tn_lower}`" in content:
                referenced_in_code.setdefault(tn, set()).add(f"View: {v_name}")

    print(f"-> {len(referenced_in_code)} Tables được tham chiếu trực tiếp trong Code SP/View.")

    # 4. Phân loại và tổng hợp kết quả
    table_results = []
    active_count = 0
    unused_count = 0

    for idx, t_id in enumerate(table_names, 1):
        meta = table_meta[t_id]
        num_rows = meta['total_rows']
        size_mb = round(meta['total_bytes'] / (1024 * 1024), 2)
        
        last_mod_dt = meta['last_modified']
        last_mod_str = last_mod_dt.strftime('%Y-%m-%d %H:%M:%S') if last_mod_dt else 'N/A'
        
        job_info = table_job_usage.get(t_id)
        has_job_query = job_info is not None
        last_queried_str = job_info['last_queried'].strftime('%Y-%m-%d %H:%M:%S') if has_job_query else 'Không query trong 180d'
        users_str = job_info['user_emails'] if has_job_query else 'Không có'
        
        code_refs = referenced_in_code.get(t_id, set())
        has_code_ref = len(code_refs) > 0
        code_ref_str = ", ".join(sorted(list(code_refs))) if has_code_ref else 'Không có'
        
        # Đánh giá trạng thái Active vs Unused
        is_active = has_job_query or has_code_ref
        
        sources = []
        if has_code_ref:
            sources.append("Trong Code SP/View")
        if has_job_query:
            sources.append("Query history 180d")
        source_str = " + ".join(sources) if sources else "Không có"
        
        if is_active:
            active_count += 1
            status = 'Active (Đang sử dụng)'
        else:
            unused_count += 1
            status = 'Lâu chưa query / Unused'
            
        table_results.append({
            'STT': idx,
            'Tên Table': t_id,
            'Dataset': DATASET_ID,
            'Trạng thái': status,
            'Nguồn ghi nhận Usage': source_str,
            'Số dòng (Rows)': num_rows,
            'Dung lượng (MB)': size_mb,
            'Cập nhật dữ liệu lần cuối (Last Modified)': last_mod_str,
            'Lần cuối Query (JOBS 180d)': last_queried_str,
            'Tài khoản/BI Tool query': users_str,
            'Tham chiếu trong Code (SP/View)': 'Có' if has_code_ref else 'Không',
            'Chi tiết Code tham chiếu': code_ref_str
        })

    df = pd.DataFrame(table_results)
    
    print("\n================ KẾT QUẢ TỔNG HỢP STAGING TABLES ================")
    print(f"Tổng số Tables trong dataset 'staging': {len(df)}")
    print(f"  - Tables ĐANG SỬ DỤNG (Active): {active_count}")
    print(f"  - Tables LÂU CHƯA QUERY / UNUSED: {unused_count}")
    print("=================================================================")

    # 5. Xuất Excel định dạng chuẩn đẹp
    excel_path = EXCEL_REPORT_PATH
    try:
        if os.path.exists(excel_path):
            os.remove(excel_path)
    except Exception:
        excel_path = r'd:\bigquery\danh_sach_table_staging_usage_moi.xlsx'

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Staging Tables Usage')
        
        workbook = writer.book
        worksheet = writer.sheets['Staging Tables Usage']
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Navy Blue
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        worksheet.row_dimensions[1].height = 28
        
        data_font = Font(name='Segoe UI', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            use_zebra = (row_num % 2 == 1)
            
            is_active_row = df.iloc[row_num - 2]['Trạng thái'].startswith('Active')
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if use_zebra:
                    cell.fill = zebra_fill
                    
                if col_num in [1, 3, 5, 8, 9, 10, 11]:
                    cell.alignment = center_align
                elif col_num in [6, 7]: # Rows, Size MB
                    cell.alignment = right_align
                elif col_num == 4: # Trang thai
                    cell.alignment = center_align
                    if is_active_row:
                        cell.font = Font(name='Segoe UI', size=10, color='385723', bold=True) # Green
                    else:
                        cell.font = Font(name='Segoe UI', size=10, color='C00000', italic=True) # Red
                else:
                    cell.alignment = left_align
                    
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    print(f"\n[+] Đã xuất báo cáo Excel kiểm tra STAGING TABLES tại: {excel_path}")

if __name__ == '__main__':
    analyze_staging_tables()
