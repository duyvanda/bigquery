import os
import glob
import re
import sys
import pandas as pd
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

CREDENTIALS_PATH = 'D:/bigquery1508.json'
PROJECT_ID = 'spatial-vision-343005'
DATASET_ID = 'staging'
EXCEL_REPORT_PATH = r'd:\bigquery\danh_sach_table_staging_usage.xlsx'
ACTIVE_JOB_FILE = r'd:\bigquery\sp_handle\call_active_job.md'

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = CREDENTIALS_PATH
client = bigquery.Client(project=PROJECT_ID)

# RULE CHECK ACTIVES:
# Quét tham chiếu trực tiếp trong:
# - staging_temp/active/*.sql (107 Active SPs)
# - warehouse_view/active/*.sql (212 Active Views)
# - INFORMATION_SCHEMA.JOBS (lịch sử query 180 ngày)

def fetch_table_info(item, table_job_usage, referenced_in_code):
    try:
        t_obj = client.get_table(item.reference)
        t_id = t_obj.table_id
        
        created_str = t_obj.created.strftime('%Y-%m-%d %H:%M:%S') if t_obj.created else ''
        modified_str = t_obj.modified.strftime('%Y-%m-%d %H:%M:%S') if t_obj.modified else ''
        
        num_rows = t_obj.num_rows if t_obj.num_rows is not None else 0
        size_mb = round((t_obj.num_bytes or 0) / (1024 * 1024), 2)
        
        job_info = table_job_usage.get(t_id)
        has_job_query = job_info is not None
        last_queried_str = job_info['last_queried'].strftime('%Y-%m-%d %H:%M:%S') if has_job_query else 'Không query trong 180d'
        users_str = job_info['user_emails'] if has_job_query else 'Không có'
        
        code_refs = referenced_in_code.get(t_id, set())
        has_code_ref = len(code_refs) > 0
        code_ref_str = ", ".join(sorted(list(code_refs))) if has_code_ref else 'Không có'
        
        # RULE 3 TIÊU CHÍ
        is_active = has_job_query or has_code_ref
        
        sources = []
        if has_job_query:
            sources.append("Query history (180d)")
        if has_code_ref:
            sp_count = sum(1 for r in code_refs if r.startswith("Active SP"))
            view_count = sum(1 for r in code_refs if r.startswith("Active View"))
            if sp_count > 0:
                sources.append(f"Active SP ({sp_count})")
            if view_count > 0:
                sources.append(f"Active View ({view_count})")
                
        source_str = " + ".join(sources) if sources else "Hoàn toàn không sử dụng"
        status = 'Active (Đang sử dụng)' if is_active else 'Lâu chưa query / Unused'
            
        return {
            'Tên Table': t_id,
            'Dataset': DATASET_ID,
            'Trạng thái': status,
            'Nguồn ghi nhận Usage': source_str,
            'Số dòng (Rows)': num_rows,
            'Dung lượng (MB)': size_mb,
            'Ngày tạo (Created)': created_str,
            'Lần cuối cập nhật dữ liệu (Modified)': modified_str,
            'Lần cuối Query (JOBS 180d)': last_queried_str,
            'Tài khoản/BI Tool query': users_str,
            'Tham chiếu trong Code (Active SP/View)': 'Có' if has_code_ref else 'Không',
            'Chi tiết Code tham chiếu': code_ref_str
        }
    except Exception as e:
        return None

def analyze_staging_tables():
    print("==========================================")
    print("PHÂN TÍCH CHUẨN XÁC USAGE TABLES DATASET STAGING (TỪ THƯ MỤC ACTIVE)")
    print("==========================================\n")

    # 1. Lấy danh sách tất cả Base Tables trong dataset 'staging'
    print(f"1. Đang lấy danh sách Base Tables từ BigQuery Dataset '{DATASET_ID}'...")
    all_items = list(client.list_tables(DATASET_ID))
    base_tables = [item for item in all_items if item.table_type == 'TABLE']
    table_names = [t.table_id for t in base_tables]
    print(f"-> Tìm thấy {len(base_tables)} Base Tables trong dataset '{DATASET_ID}'.")

    # 2. Lịch sử Query (JOBS 180 ngày) - BỎ QUA SELECT * SAO LƯU
    print("\n2. Đang đọc lịch sử Query (180d) từ BigQuery JOBS...")
    query_jobs = f"""
    SELECT 
        REGEXP_EXTRACT(query, r'(?i)(?:{DATASET_ID}|`{DATASET_ID}`)\.(`?[\w]+`?)') AS raw_tbl,
        MAX(creation_time) AS last_queried_time,
        COUNT(1) AS total_queries,
        STRING_AGG(DISTINCT user_email, ', ') AS user_emails
    FROM `region-asia-southeast1`.INFORMATION_SCHEMA.JOBS
    WHERE creation_time >= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 180 DAY)
      AND (LOWER(query) LIKE '%{DATASET_ID}%' OR LOWER(query) LIKE '%table%')
      AND NOT (LOWER(query) LIKE 'select * from%' AND user_email = 'bigquery@spatial-vision-343005.iam.gserviceaccount.com')
    GROUP BY raw_tbl
    HAVING raw_tbl IS NOT NULL
    """
    
    table_job_usage = {}
    try:
        for row in client.query(query_jobs).result():
            clean_tbl = str(row.raw_tbl).replace('`', '').strip()
            table_job_usage[clean_tbl] = {
                'last_queried': row.last_queried_time,
                'total_queries': row.total_queries,
                'user_emails': row.user_emails or ''
            }
        print(f"-> Ghi nhận {len(table_job_usage)} Tables có lịch sử truy vấn thực sự trong 180 ngày.")
    except Exception as e:
        print(f"[!] Lỗi truy vấn JOBS: {e}")

    # 3. Quét tham chiếu TRONG CÁC FOLDER ACTIVE
    print("\n3. Đang quét tham chiếu Table trong staging_temp/active (107 SPs) & warehouse_view/active (212 Views)...")
    sp_files = glob.glob(r'd:\bigquery\staging_temp\active\*.sql')
    view_files = glob.glob(r'd:\bigquery\warehouse_view\active\*.sql')
    
    referenced_in_code = {}

    for sf in sp_files:
        sp_name = os.path.basename(sf)
        with open(sf, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
        for tn in table_names:
            tn_lower = tn.lower()
            if f"staging.{tn_lower}" in content or f"`staging`.`{tn_lower}`" in content or f"`{tn_lower}`" in content:
                referenced_in_code.setdefault(tn, set()).add(f"Active SP: {sp_name}")

    for vf in view_files:
        v_name = os.path.basename(vf)
        with open(vf, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
        for tn in table_names:
            tn_lower = tn.lower()
            if f"staging.{tn_lower}" in content or f"`staging`.`{tn_lower}`" in content:
                referenced_in_code.setdefault(tn, set()).add(f"Active View: {v_name}")

    print(f"-> {len(referenced_in_code)} Tables được tham chiếu trực tiếp trong Active SP/View.")

    # 4. Đọc thông tin song song 30 Threads
    print("\n4. Đang tải song song Metadata 30 threads...")
    table_results = []
    
    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = [executor.submit(fetch_table_info, item, table_job_usage, referenced_in_code) for item in base_tables]
        for future in as_completed(futures):
            res = future.result()
            if res:
                table_results.append(res)

    df = pd.DataFrame(table_results)
    df.sort_values(by='Tên Table', inplace=True)
    df.insert(0, 'STT', range(1, len(df) + 1))

    active_count = len(df[df['Trạng thái'].str.startswith('Active')])
    unused_count = len(df) - active_count
    
    print("\n================ KẾT QUẢ TỔNG HỢP STAGING TABLES CHUẨN XÁC ================")
    print(f"Tổng số Base Tables còn lại trong 'staging': {len(df)}")
    print(f"  - Tables ĐANG SỬ DỤNG TRỰC TIẾP (Active): {active_count}")
    print(f"  - Tables LÂU CHƯA QUERY / UNUSED THẬT SỰ: {unused_count}")
    print("===========================================================================")

    # 5. Xuất Excel
    excel_path = EXCEL_REPORT_PATH
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Staging Tables Usage')
        
        workbook = writer.book
        worksheet = writer.sheets['Staging Tables Usage']
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        worksheet.row_dimensions[1].height = 28
        
        data_font = Font(name='Segoe UI', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            use_zebra = (row_num % 2 == 1)
            
            is_active_row = df.iloc[row_num - 2]['Trạng thái'].startswith('Active')
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if use_zebra:
                    cell.fill = zebra_fill
                    
                if col_num in [1, 3, 5, 7, 8, 9, 10, 11]:
                    cell.alignment = center_align
                elif col_num in [5, 6]:
                    cell.alignment = right_align
                elif col_num == 4:
                    cell.alignment = center_align
                    if is_active_row:
                        cell.font = Font(name='Segoe UI', size=10, color='385723', bold=True)
                    else:
                        cell.font = Font(name='Segoe UI', size=10, color='C00000', italic=True)
                else:
                    cell.alignment = left_align
                    
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

    print(f"\n[+] Đã xuất báo cáo Excel chuẩn từ thư mục active tại: {EXCEL_REPORT_PATH}")

if __name__ == '__main__':
    analyze_staging_tables()
