import os
import glob
import re
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

CREDENTIALS_PATH = 'D:/bigquery1508.json'
PROJECT_ID = 'spatial-vision-343005'
DATASET_ID = 'warehouse'
EXCEL_REPORT_PATH = r'd:\bigquery\danh_sach_table_warehouse_usage.xlsx'

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = CREDENTIALS_PATH
client = bigquery.Client(project=PROJECT_ID)

def analyze_warehouse_tables():
    print("==========================================")
    print("PHÂN TÍCH USAGE BASE TABLES TRONG DATASET WAREHOUSE")
    print("==========================================\n")

    # 1. Lấy danh sách BASE TABLES (loại bỏ VIEWs)
    print(f"1. Đang lọc danh sách Base Tables trong dataset '{DATASET_ID}'...")
    all_items = list(client.list_tables(DATASET_ID))
    base_tables = [item for item in all_items if item.table_type == 'TABLE']
    print(f"-> Tìm thấy {len(base_tables)} Base Tables (bỏ qua {len(all_items) - len(base_tables)} Views).")

    # 2. Lấy lịch sử JOBS (180d)
    print("\n2. Đang đọc lịch sử Query (180d) từ BigQuery JOBS...")
    query_jobs = f"""
    SELECT 
        REGEXP_EXTRACT(query, r'(?i)(?:{DATASET_ID}|`{DATASET_ID}`)\.(`?[\w]+`?)') AS raw_tbl,
        MAX(creation_time) AS last_queried_time,
        COUNT(1) AS total_queries,
        STRING_AGG(DISTINCT user_email, ', ') AS user_emails
    FROM `region-asia-southeast1`.INFORMATION_SCHEMA.JOBS
    WHERE creation_time >= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 180 DAY)
      AND (LOWER(query) LIKE '%{DATASET_ID}%' OR LOWER(query) LIKE '%table%')
    GROUP BY raw_tbl
    HAVING raw_tbl IS NOT NULL
    """
    
    table_job_usage = {}
    try:
        for row in client.query(query_jobs).result():
            clean_tbl = str(row.raw_tbl).replace('`', '').strip()
            table_job_usage[clean_tbl] = {
                'last_queried': row.last_queried_time,
                'total_queries': row.total_queries,
                'user_emails': row.user_emails or ''
            }
        print(f"-> Ghi nhận {len(table_job_usage)} Tables có lịch sử truy vấn trong 180 ngày.")
    except Exception as e:
        print(f"[!] Lỗi truy vấn JOBS: {e}")

    # 3. Quét tham chiếu trong SPs và Views
    print("\n3. Đang quét tham chiếu Table trong Code SPs (staging_temp) & Views (warehouse_view)...")
    sp_files = glob.glob(r'd:\bigquery\staging_temp\*.sql')
    view_files = glob.glob(r'd:\bigquery\warehouse_view\*.sql')
    
    table_names = [t.table_id for t in base_tables]
    referenced_in_code = {}

    for sf in sp_files:
        sp_name = os.path.basename(sf)
        with open(sf, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
        for tn in table_names:
            tn_lower = tn.lower()
            if f"warehouse.{tn_lower}" in content or f"`warehouse`.`{tn_lower}`" in content or f"`{tn_lower}`" in content:
                referenced_in_code.setdefault(tn, set()).add(f"SP: {sp_name}")

    for vf in view_files:
        v_name = os.path.basename(vf)
        with open(vf, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
        for tn in table_names:
            tn_lower = tn.lower()
            if f"warehouse.{tn_lower}" in content or f"`warehouse`.`{tn_lower}`" in content:
                referenced_in_code.setdefault(tn, set()).add(f"View: {v_name}")

    print(f"-> {len(referenced_in_code)} Base Tables được tham chiếu trực tiếp trong Code SP/View.")

    # 4. Lấy chi tiết Metadata của từng Base Table
    print("\n4. Đang tải Metadata (Created, Modified, Rows, Size) cho từng Base Table...")
    table_results = []
    
    for idx, item in enumerate(base_tables, 1):
        if idx % 30 == 0 or idx == len(base_tables):
            print(f"   └─ Đã xử lý {idx}/{len(base_tables)} tables...")
            
        try:
            t_obj = client.get_table(item.reference)
            t_id = t_obj.table_id
            
            created_str = t_obj.created.strftime('%Y-%m-%d %H:%M:%S') if t_obj.created else ''
            modified_str = t_obj.modified.strftime('%Y-%m-%d %H:%M:%S') if t_obj.modified else ''
            
            num_rows = t_obj.num_rows if t_obj.num_rows is not None else 0
            size_mb = round((t_obj.num_bytes or 0) / (1024 * 1024), 2)
            
            job_info = table_job_usage.get(t_id)
            has_job_query = job_info is not None
            last_queried_str = job_info['last_queried'].strftime('%Y-%m-%d %H:%M:%S') if has_job_query else 'Không query trong 180d'
            users_str = job_info['user_emails'] if has_job_query else 'Không có'
            
            code_refs = referenced_in_code.get(t_id, set())
            has_code_ref = len(code_refs) > 0
            code_ref_str = ", ".join(sorted(list(code_refs))) if has_code_ref else 'Không có'
            
            is_active = has_job_query or has_code_ref
            
            sources = []
            if has_code_ref:
                sources.append("Trong Code SP/View")
            if has_job_query:
                sources.append("Query history 180d")
            source_str = " + ".join(sources) if sources else "Không có"
            
            if is_active:
                status = 'Active (Đang sử dụng)'
            else:
                status = 'Lâu chưa query / Unused'
                
            table_results.append({
                'STT': idx,
                'Tên Base Table': t_id,
                'Dataset': DATASET_ID,
                'Trạng thái': status,
                'Nguồn ghi nhận Usage': source_str,
                'Số dòng (Rows)': num_rows,
                'Dung lượng (MB)': size_mb,
                'Ngày tạo (Created)': created_str,
                'Lần cuối cập nhật dữ liệu (Modified)': modified_str,
                'Lần cuối Query (JOBS 180d)': last_queried_str,
                'Tài khoản/BI Tool query': users_str,
                'Tham chiếu trong Code (SP/View)': 'Có' if has_code_ref else 'Không',
                'Chi tiết Code tham chiếu': code_ref_str
            })
            
        except Exception as e:
            print(f"Lỗi lấy thông tin table {item.table_id}: {e}")

    df = pd.DataFrame(table_results)
    
    active_count = len(df[df['Trạng thái'].str.startswith('Active')])
    unused_count = len(df) - active_count
    
    print("\n================ KẾT QUẢ TỔNG HỢP WAREHOUSE BASE TABLES ================")
    print(f"Tổng số Base Tables trong dataset 'warehouse': {len(df)}")
    print(f"  - Base Tables ĐANG SỬ DỤNG (Active): {active_count}")
    print(f"  - Base Tables LÂU CHƯA QUERY / UNUSED: {unused_count}")
    print("=========================================================================")

    # 5. Xuất Excel định dạng chuẩn đẹp
    excel_path = EXCEL_REPORT_PATH
    try:
        if os.path.exists(excel_path):
            os.remove(excel_path)
    except Exception:
        excel_path = r'd:\bigquery\danh_sach_table_warehouse_usage_moi.xlsx'

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Warehouse Tables Usage')
        
        workbook = writer.book
        worksheet = writer.sheets['Warehouse Tables Usage']
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Navy Blue
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        worksheet.row_dimensions[1].height = 28
        
        data_font = Font(name='Segoe UI', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            use_zebra = (row_num % 2 == 1)
            
            is_active_row = df.iloc[row_num - 2]['Trạng thái'].startswith('Active')
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if use_zebra:
                    cell.fill = zebra_fill
                    
                if col_num in [1, 3, 5, 8, 9, 10, 11, 12]:
                    cell.alignment = center_align
                elif col_num in [6, 7]: # Rows, Size MB
                    cell.alignment = right_align
                elif col_num == 4: # Trang thai
                    cell.alignment = center_align
                    if is_active_row:
                        cell.font = Font(name='Segoe UI', size=10, color='385723', bold=True) # Green
                    else:
                        cell.font = Font(name='Segoe UI', size=10, color='C00000', italic=True) # Red
                else:
                    cell.alignment = left_align
                    
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    print(f"\n[+] Đã xuất báo cáo Excel kiểm tra WAREHOUSE TABLES tại: {excel_path}")

if __name__ == '__main__':
    analyze_warehouse_tables()
