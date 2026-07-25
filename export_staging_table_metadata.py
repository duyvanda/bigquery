import os
import sys
import pandas as pd
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

CREDENTIALS_PATH = 'D:/bigquery1508.json'
PROJECT_ID = 'spatial-vision-343005'
DATASET_ID = 'staging'
EXCEL_OUTPUT_PATH = r'd:\bigquery\report_staging_tables_metadata.xlsx'

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = CREDENTIALS_PATH
client = bigquery.Client(project=PROJECT_ID)

def fetch_single_table_meta(item):
    try:
        t_obj = client.get_table(item.reference)
        t_id = t_obj.table_id
        
        created_dt = t_obj.created
        modified_dt = t_obj.modified
        
        created_str = created_dt.strftime('%Y-%m-%d %H:%M:%S') if created_dt else 'N/A'
        modified_str = modified_dt.strftime('%Y-%m-%d %H:%M:%S') if modified_dt else 'N/A'
        
        num_rows = t_obj.num_rows if t_obj.num_rows is not None else 0
        num_bytes = t_obj.num_bytes or 0
        
        size_mb = round(num_bytes / (1024 * 1024), 2)
        size_gb = round(num_bytes / (1024 * 1024 * 1024), 4)
        
        return {
            'Tên Table': t_id,
            'Dataset': DATASET_ID,
            'Loại': 'BASE TABLE',
            'Số dòng (Rows)': num_rows,
            'Dung lượng (MB)': size_mb,
            'Dung lượng (GB)': size_gb,
            'Ngày tạo (Created)': created_str,
            'Ngày cập nhật cuối (Last Modified)': modified_str
        }
    except Exception as e:
        return None

def export_staging_metadata():
    print("==========================================")
    print("XUẤT BÁO CÁO METADATA TABLES STAGING (DÙNG THREAD POOL CỰC NHANH)")
    print("==========================================\n")

    print(f"1. Đang lấy danh sách Tables từ Dataset '{PROJECT_ID}.{DATASET_ID}'...")
    all_items = list(client.list_tables(DATASET_ID))
    base_tables = [item for item in all_items if item.table_type == 'TABLE']
    print(f"-> Tìm thấy {len(base_tables)} Base Tables (bỏ qua {len(all_items) - len(base_tables)} Views).\n")

    print("2. Đang tải song song Metadata 30 threads...")
    results = []
    
    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = [executor.submit(fetch_single_table_meta, item) for item in base_tables]
        for future in as_completed(futures):
            res = future.result()
            if res:
                results.append(res)

    df = pd.DataFrame(results)

    # Sắp xếp theo ngày cập nhật cuối giảm dần
    df.sort_values(by='Ngày cập nhật cuối (Last Modified)', ascending=False, inplace=True)
    df.insert(0, 'STT', range(1, len(df) + 1))

    print("\n================ TỔNG HỢP METADATA STAGING TABLES ================")
    print(f"Tổng số Base Tables: {len(df)}")
    print(f"Tổng dung lượng: {round(df['Dung lượng (MB)'].sum(), 2):,} MB ({round(df['Dung lượng (GB)'].sum(), 2):,} GB)")
    print(f"Tổng số dòng dữ liệu: {df['Số dòng (Rows)'].sum():,} rows")
    print("===================================================================")

    # 3. Xuất file Excel định dạng chuẩn đẹp
    print(f"\n3. Đang ghi file Excel tại: {EXCEL_OUTPUT_PATH}...")
    
    with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Staging Tables Metadata')
        
        workbook = writer.book
        worksheet = writer.sheets['Staging Tables Metadata']
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Navy Blue
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        worksheet.row_dimensions[1].height = 28
        
        data_font = Font(name='Segoe UI', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            use_zebra = (row_num % 2 == 1)
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if use_zebra:
                    cell.fill = zebra_fill
                    
                if col_num in [1, 3, 4, 8, 9]:
                    cell.alignment = center_align
                elif col_num in [5, 6, 7]:
                    cell.alignment = right_align
                    if col_num == 5:
                        cell.number_format = '#,##0'
                    elif col_num in [6, 7]:
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = left_align
                    
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

    print(f"\n[+] XUẤT BÁO CÁO THÀNH CÔNG TẠI: {EXCEL_OUTPUT_PATH}")

if __name__ == '__main__':
    export_staging_metadata()
