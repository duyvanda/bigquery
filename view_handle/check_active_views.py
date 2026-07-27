import os
import glob
import re
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

VIEW_CACHE_PATH = r'd:\bigquery\cache\view_jobs_cache.csv'

CREDENTIALS_PATH = 'D:/bigquery1508.json'
PROJECT_ID = 'spatial-vision-343005'
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = CREDENTIALS_PATH

client = bigquery.Client(project=PROJECT_ID) if not os.path.exists(VIEW_CACHE_PATH) else None

def analyze_view_usage():
    print("==========================================")
    print("PHÂN TÍCH USAGE VIEWS (LOOKER / BI / CODE SP) - CLIENT-SIDE SCAN")
    print("==========================================\n")

    # 1. Lấy danh sách views từ active/ + backup/ (đã phân folder)
    active_view_files = glob.glob(r'd:\bigquery\warehouse_view\active\*.sql')
    backup_view_files = glob.glob(r'd:\bigquery\warehouse_view\backup\*.sql')
    view_files = active_view_files + backup_view_files
    view_names = sorted(list(set(os.path.splitext(os.path.basename(f))[0].strip() for f in view_files)))
    print(f"1. Tổng số VIEWs local (active + backup): {len(view_names)} ({len(active_view_files)} active, {len(backup_view_files)} backup)")

    # 2. Quét tham chiếu VIEW trong các Stored Procedure active (.sql) & VIEW active khác
    sp_files = glob.glob(r'd:\bigquery\staging_temp\active\*.sql')
    referenced_in_code = {}

    print(f"2. Đang quét tham chiếu trong {len(sp_files)} Active SPs và {len(active_view_files)} Active VIEWs...")

    for sf in sp_files:
        sp_name = os.path.basename(sf)
        with open(sf, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
        for vn in view_names:
            vn_lower = vn.lower()
            if f"warehouse.{vn_lower}" in content or f"`warehouse`.`{vn_lower}`" in content or f"`{vn_lower}`" in content:
                referenced_in_code.setdefault(vn, set()).add(f"SP: {sp_name}")

    for vf in active_view_files:
        v_name = os.path.splitext(os.path.basename(vf))[0]
        with open(vf, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
        for vn in view_names:
            if vn == v_name:
                continue
            vn_lower = vn.lower()
            if f"warehouse.{vn_lower}" in content or f"`warehouse`.`{vn_lower}`" in content:
                referenced_in_code.setdefault(vn, set()).add(f"View: {v_name}")

    print(f"   -> {len(referenced_in_code)} VIEWs có tham chiếu trong Code SP / View DDL.")

    # 3. Đọc View usage từ cache local (chạy cache\fetch_jobs_cache.py trước nếu chưa có)
    print("\n3. Đọc View JOBS cache từ local...")

    view_job_usage = {}  # view_name -> {'last_used': ..., 'users': ...}

    if os.path.exists(VIEW_CACHE_PATH):
        print(f"   -> Cache tìm thấy: {VIEW_CACHE_PATH}")
        df_cache = pd.read_csv(VIEW_CACHE_PATH, parse_dates=['last_used_time'])
        # Filter lại 180 ngày tính từ NOW (cache có thể cũ hơn ngày fetch)
        cutoff = pd.Timestamp.now(tz='UTC') - pd.Timedelta(days=180)
        df_cache['last_used_time'] = pd.to_datetime(df_cache['last_used_time'], utc=True)
        df_cache = df_cache[df_cache['last_used_time'] >= cutoff]
        print(f"   -> Cache có {len(df_cache)} view references trong 180 ngày gần nhất.")
        for _, row in df_cache.iterrows():
            vn = str(row['name']).strip().lower()
            orig = next((v for v in view_names if v.lower() == vn), None)
            if orig:
                existing = view_job_usage.get(orig)
                t = row['last_used_time']
                if existing is None or t > existing['last_used']:
                    view_job_usage[orig] = {
                        'last_used': t,
                        'users': str(row['users']) if pd.notna(row['users']) else ''
                    }
    else:
        print(f"   -> Cache chưa có, đang query BigQuery (chạy cache/fetch_jobs_cache.py để tạo cache)...")
        query_jobs = """
        SELECT
            LOWER(REPLACE(view_name, '`', '')) AS view_name,
            MAX(creation_time)                 AS last_used_time,
            STRING_AGG(DISTINCT user_email, ', ') AS user_emails
        FROM `region-asia-southeast1`.INFORMATION_SCHEMA.JOBS,
        UNNEST(REGEXP_EXTRACT_ALL(
            LOWER(query),
            r'(?:warehouse|`warehouse`)[.`]+([a-z0-9_]+)'
        )) AS view_name
        WHERE creation_time >= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 180 DAY)
        GROUP BY view_name
        """
        try:
            job_rows = list(client.query(query_jobs).result())
            print(f"   -> BigQuery trả về {len(job_rows)} view references.")
            for row in job_rows:
                vn = str(row.view_name).strip().lower()
                orig = next((v for v in view_names if v.lower() == vn), None)
                if orig:
                    existing = view_job_usage.get(orig)
                    if existing is None or row.last_used_time > existing['last_used']:
                        view_job_usage[orig] = {
                            'last_used': row.last_used_time,
                            'users': row.user_emails or ''
                        }
        except Exception as e:
            print(f"   [!] Lỗi truy vấn JOBS: {e}")

    print(f"   -> Ghi nhận {len(view_job_usage)} VIEWs có lịch sử truy vấn trên BigQuery!")
                            'last_used': row.last_used_time,
                            'users': row.user_emails or ''
                        }
        print(f"   -> Ghi nhận {len(view_job_usage)} VIEWs có lịch sử truy vấn trên BigQuery!")
    except Exception as e:
        print(f"   [!] Lỗi truy vấn JOBS: {e}")

    # 4. Phân loại tổng hợp
    summary_list = []
    active_count = 0
    unused_count = 0

    for idx, vn in enumerate(view_names, 1):
        refs = referenced_in_code.get(vn, set())
        is_code_ref = len(refs) > 0
        
        job_info = view_job_usage.get(vn)
        is_job_ref = job_info is not None
        
        # ACTIVE NẾU CÓ TRONG CODE HOẶC TRONG LOOKER/BI JOBS
        is_active = is_code_ref or is_job_ref
        
        last_used_str = job_info['last_used'].strftime('%Y-%m-%d %H:%M:%S') if is_job_ref else 'Không chạy trong 180 ngày qua'
        user_list_str = job_info['users'] if is_job_ref else 'Không có'
        
        sources = []
        if is_code_ref:
            sources.append("Trong Code SP/View")
        if is_job_ref:
            sources.append("Looker/BI/Direct Query")
            
        source_str = " + ".join(sources) if sources else "Không có"
        
        if is_active:
            active_count += 1
            status = 'Active (Đang sử dụng)'
        else:
            unused_count += 1
            status = 'Unused (Không dùng)'
            
        ref_details = ", ".join(sorted(list(refs))) if refs else 'Không có'
        
        summary_list.append({
            'STT': idx,
            'Tên VIEW': vn,
            'Dataset': 'warehouse',
            'Trạng thái': status,
            'Nguồn sử dụng': source_str,
            'Lần cuối truy vấn (Looker/Jobs 180d)': last_used_str,
            'Tài khoản/BI Tool truy vấn': user_list_str,
            'Chi tiết Tham chiếu trong SP/View': ref_details
        })

    df = pd.DataFrame(summary_list)
    
    print("\n================ KẾT QUẢ PHÂN TÍCH TỔNG HỢP (LOOKER + CODE) ================")
    print(f"Tổng số VIEWs kiểm tra: {len(view_names)}")
    print(f"  - VIEWs ĐANG SỬ DỤNG (Active): {active_count}")
    print(f"  - VIEWs KHÔNG SỬ DỤNG (Unused): {unused_count}")
    print("==========================================================================")

    # Export Excel với format đẹp
    excel_path = r'd:\bigquery\danh_sach_view_usage.xlsx'
    
    try:
        if os.path.exists(excel_path):
            os.remove(excel_path)
    except Exception:
        excel_path = r'd:\bigquery\danh_sach_view_usage_moi.xlsx'

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Kiểm Tra View Usage')
        
        workbook = writer.book
        worksheet = writer.sheets['Kiểm Tra View Usage']
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Navy Blue
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        worksheet.row_dimensions[1].height = 28
        
        data_font = Font(name='Segoe UI', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            use_zebra = (row_num % 2 == 1)
            
            is_active_row = df.iloc[row_num - 2]['Trạng thái'].startswith('Active')
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if use_zebra:
                    cell.fill = zebra_fill
                    
                if col_num in [1, 3, 5, 6]:
                    cell.alignment = center_align
                elif col_num == 4: # Trang thai
                    cell.alignment = center_align
                    if is_active_row:
                        cell.font = Font(name='Segoe UI', size=10, color='385723', bold=True) # Green
                    else:
                        cell.font = Font(name='Segoe UI', size=10, color='C00000', italic=True) # Red
                else:
                    cell.alignment = left_align
                    
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    print(f"\n[+] Đã xuất báo cáo Excel kiểm tra VIEW đầy đủ Looker/BI tại: {excel_path}")

if __name__ == '__main__':
    analyze_view_usage()
